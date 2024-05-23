import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")



# print(wb.sheetnames) #Properties

# print(wb["Sheet1"]) #Case sensitive

# hoja_activa = wb.active




wb.create_sheet("Sheet 3") #Creamos una hoja

hoja_tres = wb["Sheet 3"] # Usamos un objeto para asignar a esa hoja

hoja_tres.title = "New Title" # Asignamos un nuevo titulo



hoja_primera = wb["Sheet1"]

# print(
#     hoja_primera.max_row,
#     hoja_primera.max_column
# )


# celda = hoja_primera["A1"]

# celda.value = "Nuevo Nombre"

# print(celda.value)


# celda2 = hoja_primera.cell(row=2, column=1)

# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
#     )



# for fila in range(1, hoja_primera.max_row + 1):
#     for columna in range(1, hoja_primera.max_column + 1):
#         celda = hoja_primera.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)



# columna = hoja_primera["A"]
# fila = hoja_primera["1"]

# # print(columna)

# print(fila)




hoja_primera.append([1 , 2, 3])
for row in hoja_primera.rows:
    print(row)


hoja_primera.delete_cols(1,  1)

wb.save("nuevo_excel.xlsx")